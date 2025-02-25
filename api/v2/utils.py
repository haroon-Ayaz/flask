from pathlib import Path
import pandas as pd
from datetime import datetime
from sqlalchemy.exc import IntegrityError
from db.models import TestPatient, NewTestPatient
from db.extensions import db

COLUMN_MAPPING = {
    "Reporting Status": "reporting_status",
    "Test Type main": "test_type_main",
    "Test Type": "test_type",
    "Waitlist Name": "waitlist_name",
    "Unitno": "unit_no",
    "SURNAME": "surname",
    "FORENAME": "forename",
    "Points": "points",
    "Requires Pre-Add": "requires_pre_add",
    "Key Code": "key_code",
    "Comment": "comment",
    "Dated": "dated",
    "Sub Type": "sub_type",
    "Referral Priority": "referral_priority",
    "Date Added to WL": "date_added_to_wl",
    "Adj WL Start": "adj_wl_start",
    "Waitlist Name 2": "waitlist_name_2",  # If you have a second "Waitlist Name" column
    "REMVL DTTM": "remvl_dttm",
    "Weeks Wait": "weeks_wait",
    "Current RTT Waits": "current_rtt_waits",
    "Indication": "indication",
    "ID": "rxkid",  # If you want to store the Excel 'ID' into 'rxkid'
    "Appointment Date": "appointment_date",
    "Appt By Date (IPM)": "appt_by_date_ipm",
    "Appt By Date (Calculated)": "appt_by_date_calculated",
    "SHORT NOTICE FLAG": "short_notice_flag",
}

def populate_test_patients_from_excel(filepath):
    # Read the Excel file
    df = pd.read_excel(filepath)
    
    # Optional: rename columns if you have exact duplicates, e.g.:
    # df.rename(columns={"Waitlist Name": "Waitlist Name 2"}, inplace=True)
    # Adjust the above if you have multiple columns with the same name.
    
    for _, row in df.iterrows():
        # Build a dictionary of model-field -> value
        record_data = {}
        
        for excel_col, model_field in COLUMN_MAPPING.items():
            # Safely retrieve the cell value from the row
            value = row.get(excel_col, None)
            
            # If the cell is NaN or empty, make it None
            if pd.isnull(value):
                value = None
            
            # Convert date-like fields if needed
            if model_field in [
                "dated", "date_added_to_wl", "adj_wl_start",
                "remvl_dttm", "appointment_date", "appt_by_date_ipm",
                "appt_by_date_calculated"
            ]:
                # Attempt to parse as a date if not None
                if value is not None:
                    try:
                        value = pd.to_datetime(value)
                    except Exception:
                        value = None  # or handle parsing error
            
            # Convert numeric fields if needed (e.g., 'points')
            if model_field == "points" and value is not None:
                try:
                    value = int(value)
                except ValueError:
                    value = None
            
            record_data[model_field] = value
        
        # If your model requires certain fields to be non-null
        # but your Excel might have them empty, handle that logic here:
        # e.g., if 'rxkid' is required, generate something or skip:
        if not record_data.get("rxkid"):
            # generate a unique fallback or skip this row
            # record_data["rxkid"] = str(uuid.uuid4())
            pass
        
        # Create the model instance
        test_patient = TestPatient(**record_data)
        
        # Add to the session
        db.session.add(test_patient)
    
    # Commit all at once
    try:
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        print("Error inserting rows:", e)

import json
from openpyxl import load_workbook
from datetime import datetime

def read_xlsx_to_array(filepath):
    workbook = load_workbook(filename=filepath, data_only=True)
    sheet = workbook.active
    count = 0
    # Get the header row (assumes first row contains headers)
    raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    
    # Filter unique headers and record their column indices
    unique_headers = []
    unique_indices = []
    for idx, header in enumerate(raw_headers):
        if header is not None and header not in unique_headers:
            unique_headers.append(header)
            unique_indices.append(idx)
    
    data = []
    # Iterate over all rows starting from row 2 and skip completely empty rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(row[idx] is None for idx in unique_indices):
            continue  # Skip rows that are completely empty in the unique columns
        
        row_data = {}
        for header, idx in zip(unique_headers, unique_indices):
            value = row[idx]
            if isinstance(value, datetime):
                value = value.isoformat()
            if isinstance(value, str):
                value = value.strip()
            if value is None or (isinstance(value, str) and value == ""):
                value = "N/A"
            row_data[header] = value
        
        count += 1
        data.append(row_data)
  
    return data, count 

def updated_data_population():
    current_dir = Path(__file__).resolve().parent.parent
    base_dir = current_dir.parent
    file_path = base_dir / "data" / "extr.xlsx"
    try:
        records, count = read_xlsx_to_array(filepath=file_path)
        
        for record in records:
            new_patient = NewTestPatient(
                external_id=record.get("ID"),
                reporting_status=record.get("Reporting Status"),
                test_type_main=record.get("Test Type main"),
                test_type=record.get("Test Type"),
                waitlist_name=record.get("Waitlist Name"),
                rxkid=record.get("Unitno"),
                surname=record.get("SURNAME"),
                forename=record.get("FORENAME"),
                points=record.get("Points"),
                requires_pre_add=record.get("Requires Pre-Add"),
                key_code=record.get("Key Code"),
                comment=record.get("Comment ").strip() if record.get("Comment ") else None,
                dated=record.get("Dated"),
                sub_type=record.get("Sub Type"),
                referral_priority=record.get("Referral Priority"),
                date_added_to_wl=record.get("Date Added to WL"),
                adj_wl_start=record.get("Adj WL Start"),
                remvl_dttm=record.get("REMVL DTTM"),
                weeks_wait=str(record.get("Weeks Wait")),
                current_rtt_waits=record.get("Current RTT Waits"),
                indication=record.get("Indication"),
                appointment_date=record.get("Appointment Date"),
                appt_by_date_ipm=record.get("Appt By Date (IPM)"),
                appt_by_date_calculated=record.get("Appt By Date  (Calculated)"),
                short_notice_flag=record.get("SHORT NOTICE FLAG")
            )
            db.session.add(new_patient)
        db.session.commit()
        return f"Successfully populated {count} records."
    except Exception as e:
        db.session.rollback()
        return f"Error Occurred: {e}"

